"""Genera archivos descargables (Excel, PDF, gráficas PNG) a partir de datos de reportes."""
import io
from datetime import datetime

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fpdf import FPDF


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _header_fill():
    return PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")


def _header_font():
    return Font(color="FFFFFF", bold=True)


def _style_header_row(ws, row: int, cols: int):
    fill = _header_fill()
    font = _header_font()
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


# ─── Excel ────────────────────────────────────────────────────────────────────

def _excel_sales(data: dict, days: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = f"Reporte de Ventas — últimos {days} días"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {_now_str()}"
    ws.append([])

    # Totales
    ws.append(["Métrica", "Valor"])
    _style_header_row(ws, ws.max_row, 2)
    total_sales   = data.get("totalSales", data.get("total_sales", 0))
    total_revenue = data.get("totalRevenue", data.get("total_revenue", 0))
    total_orders  = data.get("totalOrders", data.get("total_orders", 0))
    ws.append(["Total ventas",  total_sales])
    ws.append(["Ingresos COP",  f"${total_revenue:,.0f}" if total_revenue else "N/A"])
    ws.append(["Total órdenes", total_orders])
    ws.append([])

    # Ventas por día
    sales_by_day = data.get("salesByDay", data.get("sales_by_day", []))
    if sales_by_day:
        ws.append(["Fecha", "Ventas", "Ingresos COP"])
        _style_header_row(ws, ws.max_row, 3)
        for row in sales_by_day:
            ws.append([
                row.get("date", row.get("fecha", "")),
                row.get("sales", row.get("ventas", 0)),
                row.get("revenue", row.get("ingresos", 0)),
            ])
        ws.append([])

    # Top productos
    top_products = data.get("topProducts", data.get("top_products", []))
    if top_products:
        ws.append(["Producto", "Unidades vendidas", "Ingresos COP"])
        _style_header_row(ws, ws.max_row, 3)
        for p in top_products:
            ws.append([
                p.get("name", p.get("nombre", "")),
                p.get("quantity", p.get("cantidad", 0)),
                p.get("revenue", p.get("ingresos", 0)),
            ])

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_stock(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    ws["A1"] = "Reporte de Inventario"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {_now_str()}"
    ws.append([])

    items = data.get("items", data.get("variants", data.get("inventory", [])))
    if items:
        ws.append(["Producto", "Variante", "Stock", "Estado"])
        _style_header_row(ws, ws.max_row, 4)
        for item in items:
            qty = item.get("quantity", item.get("stock", item.get("qty", 0)))
            estado = "Sin stock" if qty == 0 else ("Crítico" if qty <= 5 else "OK")
            ws.append([
                item.get("productName", item.get("product", item.get("name", ""))),
                item.get("variantName", item.get("variant", item.get("sku", ""))),
                qty,
                estado,
            ])
    else:
        ws.append(["Sin datos de inventario disponibles"])

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_orders(data: dict, days: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Órdenes"

    ws["A1"] = f"Reporte de Órdenes — últimos {days} días"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {_now_str()}"
    ws.append([])

    orders = data.get("orders", data.get("items", []))
    if orders:
        ws.append(["ID", "Fecha", "Estado", "Total COP", "Cliente"])
        _style_header_row(ws, ws.max_row, 5)
        for o in orders:
            ws.append([
                str(o.get("id", o.get("orderId", "")))[:8] + "...",
                o.get("createdAt", o.get("date", o.get("fecha", ""))),
                o.get("status", o.get("estado", "")),
                o.get("total", o.get("totalAmount", 0)),
                o.get("customerName", o.get("customer", o.get("cliente", ""))),
            ])
    else:
        ws.append(["Sin órdenes en el período seleccionado"])

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_dashboard(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    ws["A1"] = "Dashboard General"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generado: {_now_str()}"
    ws.append([])

    ws.append(["Métrica", "Valor"])
    _style_header_row(ws, ws.max_row, 2)
    for key, value in data.items():
        if not isinstance(value, (dict, list)):
            ws.append([str(key), str(value)])

    _autofit(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── PDF ──────────────────────────────────────────────────────────────────────

class _PDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.set_fill_color(26, 26, 46)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, "Vexio — Reporte de Tienda", align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
        self.set_text_color(0, 0, 0)
        self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f"Generado: {_now_str()} — Pág. {self.page_no()}", align="C")


def _pdf_report(title: str, rows: list[tuple], days: int | None = None) -> bytes:
    pdf = _PDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    period = f" — últimos {days} días" if days else ""
    pdf.cell(0, 8, f"{title}{period}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    if not rows:
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 8, "Sin datos disponibles para el período seleccionado.")
        return bytes(pdf.output())

    # Cabecera de tabla
    headers, *data_rows = rows
    col_w = 180 // len(headers)
    pdf.set_fill_color(26, 26, 46)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for h in headers:
        pdf.cell(col_w, 7, str(h)[:20], border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", size=9)
    for i, row in enumerate(data_rows):
        fill = i % 2 == 0
        pdf.set_fill_color(240, 240, 245) if fill else pdf.set_fill_color(255, 255, 255)
        for cell in row:
            pdf.cell(col_w, 6, str(cell)[:20], border=1, fill=fill)
        pdf.ln()

    return bytes(pdf.output())


def _pdf_sales(data: dict, days: int) -> bytes:
    top = data.get("topProducts", data.get("top_products", []))
    rows = [("Producto", "Unidades", "Ingresos COP")]
    for p in top[:20]:
        rows.append((
            p.get("name", p.get("nombre", ""))[:25],
            p.get("quantity", p.get("cantidad", 0)),
            f"${p.get('revenue', p.get('ingresos', 0)):,}",
        ))
    return _pdf_report("Reporte de Ventas", rows, days)


def _pdf_stock(data: dict) -> bytes:
    items = data.get("items", data.get("variants", data.get("inventory", [])))
    rows = [("Producto", "Variante", "Stock", "Estado")]
    for item in items[:30]:
        qty = item.get("quantity", item.get("stock", 0))
        rows.append((
            item.get("productName", item.get("name", ""))[:20],
            item.get("variantName", item.get("sku", ""))[:15],
            qty,
            "Sin stock" if qty == 0 else ("Crítico" if qty <= 5 else "OK"),
        ))
    return _pdf_report("Reporte de Inventario", rows)


def _pdf_orders(data: dict, days: int) -> bytes:
    orders = data.get("orders", data.get("items", []))
    rows = [("ID", "Fecha", "Estado", "Total")]
    for o in orders[:30]:
        rows.append((
            str(o.get("id", o.get("orderId", "")))[:8],
            str(o.get("createdAt", o.get("date", "")))[:10],
            o.get("status", o.get("estado", ""))[:12],
            f"${o.get('total', o.get('totalAmount', 0)):,}",
        ))
    return _pdf_report("Reporte de Órdenes", rows, days)


def _pdf_dashboard(data: dict) -> bytes:
    rows = [("Métrica", "Valor")]
    for key, value in data.items():
        if not isinstance(value, (dict, list)):
            rows.append((str(key)[:25], str(value)[:25]))
    return _pdf_report("Dashboard General", rows)


# ─── Gráficas ─────────────────────────────────────────────────────────────────

def _chart_sales(data: dict, days: int) -> bytes:
    sales_by_day = data.get("salesByDay", data.get("sales_by_day", []))
    top_products = data.get("topProducts", data.get("top_products", []))

    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle(f"Ventas — últimos {days} días", fontsize=14, fontweight="bold")

    # Gráfica 1: ventas por día
    if sales_by_day:
        dates = [str(r.get("date", r.get("fecha", "")))[-5:] for r in sales_by_day[-14:]]
        values = [r.get("sales", r.get("ventas", 0)) for r in sales_by_day[-14:]]
        axes[0].bar(dates, values, color="#4A90D9")
        axes[0].set_title("Ventas por día")
        axes[0].set_xlabel("Fecha")
        axes[0].set_ylabel("Unidades")
        axes[0].tick_params(axis="x", rotation=45)
        axes[0].yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    else:
        axes[0].text(0.5, 0.5, "Sin datos", ha="center", va="center")
        axes[0].set_title("Ventas por día")

    # Gráfica 2: top productos
    if top_products:
        names = [p.get("name", p.get("nombre", ""))[:15] for p in top_products[:8]]
        units = [p.get("quantity", p.get("cantidad", 0)) for p in top_products[:8]]
        axes[1].barh(names, units, color="#E8A838")
        axes[1].set_title("Top productos")
        axes[1].set_xlabel("Unidades vendidas")
        axes[1].invert_yaxis()
    else:
        axes[1].text(0.5, 0.5, "Sin datos", ha="center", va="center")
        axes[1].set_title("Top productos")

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def _chart_stock(data: dict) -> bytes:
    items = data.get("items", data.get("variants", data.get("inventory", [])))
    out_of_stock = sum(1 for i in items if i.get("quantity", i.get("stock", 0)) == 0)
    critical     = sum(1 for i in items if 0 < i.get("quantity", i.get("stock", 0)) <= 5)
    ok           = len(items) - out_of_stock - critical

    fig, ax = plt.subplots(figsize=(7, 5))
    if items:
        labels = ["Sin stock", "Crítico (1-5)", "OK"]
        sizes  = [out_of_stock, critical, ok]
        colors = ["#E74C3C", "#F39C12", "#27AE60"]
        non_zero = [(l, s, c) for l, s, c in zip(labels, sizes, colors) if s > 0]
        if non_zero:
            ax.pie(
                [s for _, s, _ in non_zero],
                labels=[l for l, _, _ in non_zero],
                colors=[c for _, _, c in non_zero],
                autopct="%1.0f%%",
                startangle=90,
            )
    ax.set_title("Estado del Inventario")
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def _chart_orders(data: dict, days: int) -> bytes:
    orders = data.get("orders", data.get("items", []))
    status_count: dict = {}
    for o in orders:
        s = o.get("status", o.get("estado", "UNKNOWN"))
        status_count[s] = status_count.get(s, 0) + 1

    fig, ax = plt.subplots(figsize=(8, 5))
    fig.suptitle(f"Órdenes por estado — últimos {days} días", fontsize=13, fontweight="bold")
    if status_count:
        ax.bar(list(status_count.keys()), list(status_count.values()), color="#8E44AD")
        ax.set_ylabel("Cantidad")
        ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    else:
        ax.text(0.5, 0.5, "Sin datos", ha="center", va="center")

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


def _chart_dashboard(data: dict) -> bytes:
    metrics = {k: v for k, v in data.items() if isinstance(v, (int, float)) and v > 0}
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.suptitle("Dashboard General", fontsize=13, fontweight="bold")
    if metrics:
        ax.bar(list(metrics.keys()), list(metrics.values()), color="#2ECC71")
        ax.tick_params(axis="x", rotation=30)
    else:
        ax.text(0.5, 0.5, "Sin datos numéricos", ha="center", va="center")
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=120, bbox_inches="tight")
    plt.close(fig)
    return buf.getvalue()


# ─── Punto de entrada principal ───────────────────────────────────────────────

def generate_report(
    data: dict,
    report_type: str,
    fmt: str,
    days: int = 30,
) -> tuple[bytes, str, str]:
    """
    Devuelve (bytes, mime_type, filename).
    report_type: sales | stock | orders | dashboard
    fmt:         excel | pdf | chart
    """
    rtype = report_type.lower()
    fmt   = fmt.lower()
    date  = _now_str()

    if fmt == "excel":
        if rtype == "sales":
            data_bytes = _excel_sales(data, days)
        elif rtype == "stock":
            data_bytes = _excel_stock(data)
        elif rtype == "orders":
            data_bytes = _excel_orders(data, days)
        else:
            data_bytes = _excel_dashboard(data)
        mime     = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"reporte_{rtype}_{date}.xlsx"

    elif fmt == "pdf":
        if rtype == "sales":
            data_bytes = _pdf_sales(data, days)
        elif rtype == "stock":
            data_bytes = _pdf_stock(data)
        elif rtype == "orders":
            data_bytes = _pdf_orders(data, days)
        else:
            data_bytes = _pdf_dashboard(data)
        mime     = "application/pdf"
        filename = f"reporte_{rtype}_{date}.pdf"

    else:  # chart / png
        if rtype == "sales":
            data_bytes = _chart_sales(data, days)
        elif rtype == "stock":
            data_bytes = _chart_stock(data)
        elif rtype == "orders":
            data_bytes = _chart_orders(data, days)
        else:
            data_bytes = _chart_dashboard(data)
        mime     = "image/png"
        filename = f"grafica_{rtype}_{date}.png"

    return data_bytes, mime, filename
